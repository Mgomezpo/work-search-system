# Genera el tracker (Dashboard + Tracker) con formulas y validaciones.\n# Correr: python build_tracker.py  |  Luego recalcular con el recalc.py del skill xlsx.\nfrom openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT="Arial"; NAVY="1F3864"; BLUE="2E5395"; LIGHT="D9E1F2"; GREY="F2F2F2"; YELLOW="FFF2CC"; WHITE="FFFFFF"
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook(); ws=wb.active; ws.title="Tracker"
headers=[("ID",6),("Fecha agregado",14),("Empresa",22),("Cargo / Titulo",34),("Link (aplicar)",18),
("Ubicacion / Modalidad",22),("Salario (USD)",20),("Salario/mes aprox (USD)",16),("Fit salario",12),
("Contexto / Descripcion",46),("Match ATS (%)",13),("CV usado",26),("Fecha aplicacion",14),
("Estado",16),("Respuesta",16),("Entrevista",28),("Notas / Proximos pasos",40)]

ws.merge_cells("A1:Q1"); t=ws["A1"]; t.value="WORK SEARCH — TRACKER DE VACANTES (PLANTILLA)"
t.font=Font(name=FONT,size=15,bold=True,color=WHITE); t.fill=PatternFill("solid",fgColor=NAVY)
t.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[1].height=28

for i,(h,w) in enumerate(headers,start=1):
    c=ws.cell(row=2,column=i,value=h)
    c.font=Font(name=FONT,size=10,bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[2].height=32

# Two illustrative EXAMPLE rows (clearly marked) so the user sees how to fill it
examples=[
 (1,"2026-01-15","EJEMPLO S.A.","Senior Data Scientist","https://ejemplo.com/aplicar","Remote",
  "$90,000 - $120,000 / yr",90000,120000,"EJEMPLO: pega aqui el contexto/descripcion de la vacante.",75,
  "CV_tailored_EjemploSA.docx","2026-01-16","Aplicado","Acuse de recibo","Por agendar","EJEMPLO: borra esta fila al empezar."),
 (2,"2026-01-18","DEMO Corp","ML Engineer","https://demo.com/aplicar","Remote LATAM",
  "$60,000 - $80,000 / yr",60000,80000,"EJEMPLO: rol de ML, stack Python+AWS.",80,
  "","","Por aplicar","Sin respuesta","—","EJEMPLO."),
]
ds=3
for k,row in enumerate(examples):
    r=ds+k
    (idv,fecha,emp,car,link,loc,sal,low,high,ctx,match,cv,fa,est,resp,intv,notas)=row
    ws.cell(r,1,idv); ws.cell(r,2,fecha); ws.cell(r,3,emp); ws.cell(r,4,car)
    lc=ws.cell(r,5,"Aplicar"); lc.hyperlink=link; lc.font=Font(name=FONT,size=10,color="0563C1",underline="single")
    ws.cell(r,6,loc); ws.cell(r,7,sal)
    ws.cell(r,8,f"=ROUND(AVERAGE({low},{high})/12,0)")
    ws.cell(r,9,f'=IF(H{r}>=6000,"Excelente",IF(H{r}>=5000,"En target",IF(H{r}>=4000,"Aceptable","Bajo")))')
    ws.cell(r,10,ctx); ws.cell(r,11,match); ws.cell(r,12,cv); ws.cell(r,13,fa)
    ws.cell(r,14,est); ws.cell(r,15,resp); ws.cell(r,16,intv); ws.cell(r,17,notas)

for r in range(ds,201):
    for col in range(1,18):
        c=ws.cell(r,col)
        if not (col==5 and r<ds+len(examples)):
            c.font=Font(name=FONT,size=10)
        c.border=border
        c.alignment=Alignment(vertical="center",wrap_text=(col in (4,10,16,17)))
        if col in (1,2,8,9,11,13,14,15):
            c.alignment=Alignment(horizontal="center",vertical="center")
    if (r-ds)%2==1:
        for col in range(1,18): ws.cell(r,col).fill=PatternFill("solid",fgColor=GREY)
    ws.row_dimensions[r].height=30

dv_status=DataValidation(type="list",formula1='"Por aplicar,Aplicado,En revision,Entrevista,Oferta,Rechazado,Descartado"',allow_blank=True)
ws.add_data_validation(dv_status); dv_status.add(f"N{ds}:N200")
dv_resp=DataValidation(type="list",formula1='"Sin respuesta,Acuse de recibo,Recruiter screen,Avanza,Rechazo,Oferta"',allow_blank=True)
ws.add_data_validation(dv_resp); dv_resp.add(f"O{ds}:O200")
dv_intv=DataValidation(type="list",formula1='"—,Por agendar,Agendada,Screen tecnico,On-site/Final,Completada,No avanzo"',allow_blank=True)
ws.add_data_validation(dv_intv); dv_intv.add(f"P{ds}:P200")
ws.freeze_panes="A3"; ws.sheet_view.showGridLines=False

dash=wb.create_sheet("Dashboard",0); dash.sheet_view.showGridLines=False
dash.merge_cells("A1:E1"); d=dash["A1"]; d.value="WORK SEARCH — DASHBOARD"
d.font=Font(name=FONT,size=18,bold=True,color=WHITE); d.fill=PatternFill("solid",fgColor=NAVY)
d.alignment=Alignment(horizontal="left",vertical="center",indent=1); dash.row_dimensions[1].height=36
dash.merge_cells("A2:E2"); sub=dash["A2"]
sub.value="Personaliza tu target arriba. Metricas automaticas. Boton Recargar si abres en Excel de escritorio."
sub.font=Font(name=FONT,size=10,italic=True,color="404040"); sub.alignment=Alignment(horizontal="left",indent=1)

metrics=[("Total vacantes",'=COUNTA(Tracker!C3:C200)'),
("Por aplicar",'=COUNTIF(Tracker!N3:N200,"Por aplicar")'),
("Aplicadas",'=COUNTIF(Tracker!N3:N200,"Aplicado")+COUNTIF(Tracker!N3:N200,"En revision")+COUNTIF(Tracker!N3:N200,"Entrevista")+COUNTIF(Tracker!N3:N200,"Oferta")+COUNTIF(Tracker!N3:N200,"Rechazado")'),
("En entrevista",'=COUNTIF(Tracker!N3:N200,"Entrevista")'),
("Ofertas",'=COUNTIF(Tracker!N3:N200,"Oferta")'),
("Rechazos",'=COUNTIF(Tracker!N3:N200,"Rechazado")')]
dash["A4"]="METRICA"; dash["B4"]="VALOR"
for cc in ("A4","B4"):
    dash[cc].font=Font(name=FONT,size=11,bold=True,color=WHITE); dash[cc].fill=PatternFill("solid",fgColor=BLUE)
    dash[cc].alignment=Alignment(horizontal="left",indent=1); dash[cc].border=border
row=5
for name,formula in metrics:
    dash.cell(row,1,name).font=Font(name=FONT,size=11)
    vc=dash.cell(row,2,formula); vc.font=Font(name=FONT,size=11,bold=True,color=NAVY); vc.alignment=Alignment(horizontal="center")
    dash.cell(row,1).border=border; dash.cell(row,2).border=border; dash.cell(row,1).fill=PatternFill("solid",fgColor=LIGHT)
    row+=1
dash.cell(row+1,1,"Tasa de respuesta").font=Font(name=FONT,size=11,bold=True)
rr=dash.cell(row+1,2,'=IFERROR((COUNTIF(Tracker!O3:O200,"Avanza")+COUNTIF(Tracker!O3:O200,"Recruiter screen")+COUNTIF(Tracker!O3:O200,"Acuse de recibo")+COUNTIF(Tracker!O3:O200,"Oferta"))/B7,"—")')
rr.number_format="0%"; rr.font=Font(name=FONT,size=11,bold=True,color=NAVY); rr.alignment=Alignment(horizontal="center")
dash.cell(row+1,1).fill=PatternFill("solid",fgColor=YELLOW)
dash.column_dimensions["A"].width=24; dash.column_dimensions["B"].width=14
dash.column_dimensions["C"].width=3; dash.column_dimensions["D"].width=46

notes=["COMO USAR","1. Pidele a Claude: 'busca trabajos en Indeed de X' -> poblara la hoja Tracker.",
"2. Cambia el Estado (lista desplegable) segun avances.","3. Registra Respuesta y Entrevista.",
"4. Pidele a Claude un CV tailored por vacante -> anota el nombre en 'CV usado'.",
"5. El Dashboard se actualiza solo.","",
"Estados: Por aplicar / Aplicado / En revision / Entrevista / Oferta / Rechazado / Descartado",
"","Borra las filas de EJEMPLO antes de empezar."]
for i,line in enumerate(notes):
    c=dash.cell(4+i,4,line)
    c.font=Font(name=FONT,size=11,bold=True,color=NAVY) if i==0 else Font(name=FONT,size=10,color="404040")

wb.save("Work_Search_Tracker.xlsx"); print("tracker saved")
